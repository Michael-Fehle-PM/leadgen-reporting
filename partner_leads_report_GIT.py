"""
partner_leads_report.py  (v5 – three tabs: MIL / FW / Rollup)
==============================================================
Produces a Partner Leads Excel workbook with three sheets:

  Tab 1 – MIL Partners  (full layout: totals + version + source breakdown)
  Tab 2 – FW Partners   (simplified: totals + CPL + scrub rate only)
  Tab 3 – Rollup        (combined totals for both BUs side by side)

HOW TO USE
----------
Option A – Oracle live connection:
    pip install cx_Oracle openpyxl
    Set DB_* constants below, then:  python partner_leads_report.py

Option B – CSV exports:
    Export MIL query to mil_results.csv and FW query to fw_results.csv, then:
    python partner_leads_report.py --csv
"""

import argparse
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Oracle connection ─────────────────────────────────────────────────────────
DB_USER     = "withheld"
DB_PASSWORD = "withheld"
DB_DSN      = "withheld"

def _output_path():
    base_dir  = os.path.dirname(os.path.abspath(__file__))
    datestamp = datetime.today().strftime("%Y%m%d")
    path = os.path.join(base_dir, f"Partner_Leads_Report_{datestamp}.xlsx")
    if not os.path.exists(path):
        return path
    counter = 1
    while True:
        path = os.path.join(base_dir, f"Partner_Leads_Report_{datestamp}_{counter}.xlsx")
        if not os.path.exists(path):
            return path
        counter += 1

OUTPUT_FILE = _output_path()

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MIL_QUERY_FILE = os.path.join(SCRIPT_DIR, "query_mil.sql")
FW_QUERY_FILE  = os.path.join(SCRIPT_DIR, "query_fw.sql")

# ── SQL column orders ─────────────────────────────────────────────────────────
MIL_COLUMNS = [
    "partner_id", "partner_name",
    "leads_mtd", "leads_prior_month", "leads_last_30",
    "cost_per_lead", "scrub_rate",
    "v1_mtd", "v1_prior_month", "v1_last_30",
    "v2_mtd", "v2_prior_month", "v2_last_30",
    "google_mtd", "google_prior_month", "google_last_30",
    "bing_mtd",   "bing_prior_month",   "bing_last_30",
    "days_elapsed_mtd", "total_days_in_month",
]

FW_COLUMNS = [
    "partner_id", "partner_name",
    "leads_mtd", "leads_prior_month", "leads_last_30",
    "cost_per_lead", "scrub_rate",
    "days_elapsed_mtd", "total_days_in_month",
]

# ── MIL column indices (1-based) ──────────────────────────────────────────────
M_PID      = 1;  M_PNAME  = 2
M_MTD      = 3;  M_PRIOR  = 4;  M_L30   = 5;  M_FCST  = 6
M_CPL      = 7;  M_SCRUB  = 8
M_V1_MTD   = 9;  M_V1_P   = 10; M_V1_L  = 11
M_V2_MTD   = 12; M_V2_P   = 13; M_V2_L  = 14
M_G_MTD    = 15; M_G_P    = 16; M_G_L   = 17; M_G_F   = 18
M_B_MTD    = 19; M_B_P    = 20; M_B_L   = 21; M_B_F   = 22
M_O_MTD    = 23; M_O_P    = 24; M_O_L   = 25; M_O_F   = 26
M_ELAPSED  = 27; M_DAYS   = 28
M_VIS_COLS = 26

# ── FW column indices (1-based) ───────────────────────────────────────────────
F_PID      = 1;  F_PNAME  = 2
F_MTD      = 3;  F_PRIOR  = 4;  F_L30   = 5;  F_FCST  = 6
F_CPL      = 7;  F_SCRUB  = 8
F_ELAPSED  = 9;  F_DAYS   = 10
F_VIS_COLS = 8

DATA_START_ROW = 6

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE    = "1F3864";  MID_BLUE    = "2E75B6";  LIGHT_BLUE   = "D6E4F0"
DARK_PURPLE  = "4B0082";  MID_PURPLE  = "7B2D8B";  LIGHT_PURPLE = "EDE7F6"
DARK_GREEN   = "375623";  MID_GREEN   = "548235";  LIGHT_GREEN  = "E2EFDA"
DARK_ORANGE  = "7B3F00";  MID_ORANGE  = "C55A11";  LIGHT_ORANGE = "FCE4D6"
ACCENT_BLUE  = "00B0F0";  ACCENT_PURP = "CE93D8"
ACCENT_GREEN = "70AD47";  ACCENT_ORG  = "F4B942"
FCST_BLUE    = "0070C0";  FCST_GREEN  = "196F3D"
WHITE        = "FFFFFF";  LIGHT_GRAY  = "F2F2F2";  YELLOW = "FFFF00"

# ── Style helpers ─────────────────────────────────────────────────────────────
def _font(bold=False, color=WHITE, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_left(left_color):
    return Border(left=Side(style="medium", color=left_color),
                  right=Side(style="thin", color="BFBFBF"),
                  top=Side(style="thin", color="BFBFBF"),
                  bottom=Side(style="thin", color="BFBFBF"))

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Data helpers ──────────────────────────────────────────────────────────────
def _n(val):
    try:    return float(val)
    except: return 0.0

def _fcst(mtd_ci, el_ci, days_ci, row_num):
    m = get_column_letter(mtd_ci)
    e = get_column_letter(el_ci)
    d = get_column_letter(days_ci)
    return f'=IFERROR(ROUND({m}{row_num}/MAX({e}{row_num},1)*{d}{row_num},0),"")'

def _org(tot_ci, g_ci, b_ci, row_num):
    t = get_column_letter(tot_ci); g = get_column_letter(g_ci); b = get_column_letter(b_ci)
    return f"=MAX({t}{row_num}-{g}{row_num}-{b}{row_num},0)"

# ── Data fetching ─────────────────────────────────────────────────────────────
def _get_conn():
    try:
        import oracledb
    except ImportError:
        sys.exit("oracledb not installed.  Run:  pip install oracledb")
    return oracledb.connect(user=DB_USER, password=DB_PASSWORD, dsn=DB_DSN)

def fetch_oracle(query_file):
    with open(query_file) as fh: sql = fh.read()
    conn = _get_conn(); cur = conn.cursor()
    cur.execute(sql); rows = cur.fetchall()
    cur.close(); conn.close()
    return rows

def fetch_csv(path, columns):
    import csv
    rows = []
    with open(path, newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append([row.get(c, 0) for c in columns])
    return rows

# ── Shared: write title + methodology row ─────────────────────────────────────
def _write_title(ws, title, note, total_cols, title_fill=DARK_BLUE):
    last = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = _font(bold=True, size=14); ws["A1"].fill = _fill(title_fill)
    ws["A1"].alignment = _center(); ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = note
    ws["A2"].font = _font(italic=True, color=DARK_BLUE, size=9)
    ws["A2"].fill = _fill(LIGHT_BLUE); ws["A2"].alignment = _left()
    ws.row_dimensions[2].height = 15

# ── Shared: write totals row ──────────────────────────────────────────────────
def _write_totals(ws, tr, last_data, sum_blue, sum_purple, sum_green):
    for ci in range(1, max(sum_blue | sum_purple | sum_green) + 1):
        cell = ws.cell(row=tr, column=ci)
        if ci in sum_green:    fill_c, acc = DARK_GREEN,  ACCENT_GREEN
        elif ci in sum_purple: fill_c, acc = DARK_PURPLE, ACCENT_PURP
        else:                  fill_c, acc = DARK_BLUE,   ACCENT_BLUE
        cell.fill = _fill(fill_c); cell.border = _border()
        if ci == 1:
            cell.value = "TOTALS"; cell.font = _font(bold=True, size=10)
            cell.alignment = _center()
        elif ci in sum_blue | sum_purple | sum_green:
            col_ltr = get_column_letter(ci)
            cell.value = f"=SUM({col_ltr}{DATA_START_ROW}:{col_ltr}{last_data})"
            cell.number_format = "#,##0"
            cell.font = _font(bold=True, color=acc, size=10); cell.alignment = _center()
    ws.row_dimensions[tr].height = 20

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 – MIL
# ─────────────────────────────────────────────────────────────────────────────
def build_mil_sheet(wb, rows):
    ws = wb.active
    ws.title = "MIL Partners"
    ws.freeze_panes = "C6"
    today = datetime.today()

    _write_title(ws,
        f"MIL Partner Leads Report  –  {today.strftime('%B %d, %Y')}",
        "Forecast: MTD ÷ Days elapsed × Days in month  |  Organic = Total − Google − Bing  |  Version: V1=default-flow  V2=default-flow-v2",
        M_VIS_COLS)

    # Row 3: section banners
    for rng, lbl, fill_c in [
        (f"A3:{get_column_letter(M_SCRUB)}3",   "◀  SECTION 1 — PARTNER PERFORMANCE", DARK_BLUE),
        (f"{get_column_letter(M_V1_MTD)}3:{get_column_letter(M_V2_L)}3", "SECTION 2 — VERSION BREAKDOWN", DARK_PURPLE),
        (f"{get_column_letter(M_G_MTD)}3:{get_column_letter(M_O_F)}3",   "SECTION 3 — LEAD SOURCE BREAKDOWN  ▶", DARK_GREEN),
    ]:
        ws.merge_cells(rng); c = ws[rng.split(":")[0]]
        c.value = lbl; c.font = _font(bold=True, size=10)
        c.fill = _fill(fill_c); c.alignment = _center()
    ws.row_dimensions[3].height = 20

    # Row 4: sub-group labels
    for ci in range(1, M_SCRUB + 1):
        ws.cell(row=4, column=ci).fill = _fill(MID_BLUE)
        ws.cell(row=4, column=ci).border = _border()
    for cols, lbl, fill_c in [
        ([M_V1_MTD, M_V1_P, M_V1_L],           "📋  Version 1",  MID_PURPLE),
        ([M_V2_MTD, M_V2_P, M_V2_L],           "📋  Version 2",  MID_PURPLE),
        ([M_G_MTD,  M_G_P,  M_G_L,  M_G_F],    "🔵  Google",     MID_BLUE),
        ([M_B_MTD,  M_B_P,  M_B_L,  M_B_F],    "🔵  Bing",       MID_BLUE),
        ([M_O_MTD,  M_O_P,  M_O_L,  M_O_F],    "🟢  Organic",    MID_GREEN),
    ]:
        sl = get_column_letter(cols[0]); el = get_column_letter(cols[-1])
        ws.merge_cells(f"{sl}4:{el}4")
        cell = ws.cell(row=4, column=cols[0], value=lbl)
        cell.font = _font(bold=True, size=10); cell.fill = _fill(fill_c)
        cell.alignment = _center(); cell.border = _border()
    ws.row_dimensions[4].height = 18

    # Row 5: column headers
    col_specs = [
        (M_PID,   "Partner\nID",              12, MID_BLUE,   WHITE,       None),
        (M_PNAME, "Partner Name",             28, MID_BLUE,   WHITE,       None),
        (M_MTD,   "Leads\nMTD",               12, MID_BLUE,   WHITE,       None),
        (M_PRIOR, "Leads\nPrior Month",       14, MID_BLUE,   WHITE,       None),
        (M_L30,   "Leads\nLast 30 Days",      14, MID_BLUE,   WHITE,       None),
        (M_FCST,  "Forecast\n(Full Month)",   14, MID_BLUE,   WHITE,       None),
        (M_CPL,   "Cost Per\nLead",           12, MID_BLUE,   WHITE,       None),
        (M_SCRUB, "Permissible\nScrub Rate",  13, MID_BLUE,   WHITE,       None),
        (M_V1_MTD,"V1\nMTD",                 11, MID_PURPLE, WHITE, DARK_PURPLE),
        (M_V1_P,  "V1\nPrior Month",         14, MID_PURPLE, WHITE,       None),
        (M_V1_L,  "V1\nLast 30 Days",        14, MID_PURPLE, WHITE,       None),
        (M_V2_MTD,"V2\nMTD",                 11, MID_PURPLE, WHITE, DARK_PURPLE),
        (M_V2_P,  "V2\nPrior Month",         14, MID_PURPLE, WHITE,       None),
        (M_V2_L,  "V2\nLast 30 Days",        14, MID_PURPLE, WHITE,       None),
        (M_G_MTD, "Google\nMTD",             12, MID_BLUE,   WHITE,  DARK_BLUE),
        (M_G_P,   "Google\nPrior Month",     14, MID_BLUE,   WHITE,       None),
        (M_G_L,   "Google\nLast 30 Days",    14, MID_BLUE,   WHITE,       None),
        (M_G_F,   "Google\nForecast",        14, MID_BLUE,   WHITE,       None),
        (M_B_MTD, "Bing\nMTD",              12, MID_BLUE,   WHITE,  DARK_BLUE),
        (M_B_P,   "Bing\nPrior Month",      14, MID_BLUE,   WHITE,       None),
        (M_B_L,   "Bing\nLast 30 Days",     14, MID_BLUE,   WHITE,       None),
        (M_B_F,   "Bing\nForecast",         14, MID_BLUE,   WHITE,       None),
        (M_O_MTD, "Organic\nMTD",           12, MID_GREEN,  WHITE, DARK_GREEN),
        (M_O_P,   "Organic\nPrior Month",   14, MID_GREEN,  WHITE,       None),
        (M_O_L,   "Organic\nLast 30 Days",  14, MID_GREEN,  WHITE,       None),
        (M_O_F,   "Organic\nForecast",      14, MID_GREEN,  WHITE,       None),
    ]
    for ci, lbl, w, fill_c, font_c, thick_c in col_specs:
        cell = ws.cell(row=5, column=ci, value=lbl)
        cell.font = _font(bold=True, color=font_c, size=9); cell.fill = _fill(fill_c)
        cell.alignment = _center()
        cell.border = _thick_left(thick_c) if thick_c else _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.cell(row=5, column=M_ELAPSED, value="Days Elapsed")
    ws.cell(row=5, column=M_DAYS,    value="Days in Month")
    ws.row_dimensions[5].height = 36

    # Data rows
    for ri, row in enumerate(rows, start=DATA_START_ROW):
        d   = {col: row[i] for i, col in enumerate(MIL_COLUMNS)}
        bg  = LIGHT_GRAY   if ri % 2 == 0 else WHITE
        pbg = LIGHT_PURPLE if ri % 2 == 0 else "F5EEF8"
        gbg = LIGHT_BLUE   if ri % 2 == 0 else "EBF3FB"
        obg = LIGHT_GREEN  if ri % 2 == 0 else "F0F7EC"

        def put(ci, value, fmt=None, fill_c=None, font_c="000000", bold=False, bdr=None):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.fill = _fill(fill_c if fill_c is not None else bg)
            cell.font = _font(color=font_c, size=10, bold=bold)
            cell.alignment = _center(); cell.border = bdr if bdr else _border()
            if fmt: cell.number_format = fmt

        put(M_PID,   d["partner_id"])
        put(M_PNAME, d["partner_name"]); ws.cell(row=ri, column=M_PNAME).alignment = _left()
        put(M_MTD,   _n(d["leads_mtd"]),         "#,##0")
        put(M_PRIOR, _n(d["leads_prior_month"]),  "#,##0")
        put(M_L30,   _n(d["leads_last_30"]),      "#,##0")
        put(M_FCST,  _fcst(M_MTD, M_ELAPSED, M_DAYS, ri), "#,##0", font_c=FCST_BLUE, bold=True)
        put(M_CPL,   _n(d["cost_per_lead"]),       '"$"#,##0.00')
        put(M_SCRUB, _n(d["scrub_rate"]),          "0.0%")
        put(M_V1_MTD, _n(d["v1_mtd"]),          "#,##0", fill_c=pbg, bdr=_thick_left(DARK_PURPLE))
        put(M_V1_P,  _n(d["v1_prior_month"]),   "#,##0", fill_c=pbg)
        put(M_V1_L,  _n(d["v1_last_30"]),       "#,##0", fill_c=pbg)
        put(M_V2_MTD, _n(d["v2_mtd"]),          "#,##0", fill_c=pbg, bdr=_thick_left(DARK_PURPLE))
        put(M_V2_P,  _n(d["v2_prior_month"]),   "#,##0", fill_c=pbg)
        put(M_V2_L,  _n(d["v2_last_30"]),       "#,##0", fill_c=pbg)
        put(M_G_MTD, _n(d["google_mtd"]),        "#,##0", fill_c=gbg, bdr=_thick_left(DARK_BLUE))
        put(M_G_P,   _n(d["google_prior_month"]), "#,##0", fill_c=gbg)
        put(M_G_L,   _n(d["google_last_30"]),    "#,##0", fill_c=gbg)
        put(M_G_F,   _fcst(M_G_MTD, M_ELAPSED, M_DAYS, ri), "#,##0", fill_c=gbg, font_c=FCST_BLUE, bold=True)
        put(M_B_MTD, _n(d["bing_mtd"]),          "#,##0", fill_c=gbg, bdr=_thick_left(DARK_BLUE))
        put(M_B_P,   _n(d["bing_prior_month"]),  "#,##0", fill_c=gbg)
        put(M_B_L,   _n(d["bing_last_30"]),      "#,##0", fill_c=gbg)
        put(M_B_F,   _fcst(M_B_MTD, M_ELAPSED, M_DAYS, ri), "#,##0", fill_c=gbg, font_c=FCST_BLUE, bold=True)
        put(M_O_MTD, _org(M_MTD,   M_G_MTD, M_B_MTD,  ri), "#,##0", fill_c=obg, font_c=DARK_GREEN, bdr=_thick_left(DARK_GREEN))
        put(M_O_P,   _org(M_PRIOR, M_G_P,   M_B_P,    ri), "#,##0", fill_c=obg, font_c=DARK_GREEN)
        put(M_O_L,   _org(M_L30,   M_G_L,   M_B_L,    ri), "#,##0", fill_c=obg, font_c=DARK_GREEN)
        put(M_O_F,   _fcst(M_O_MTD, M_ELAPSED, M_DAYS, ri), "#,##0", fill_c=obg, font_c=FCST_GREEN, bold=True)
        ws.cell(row=ri, column=M_ELAPSED).value = _n(d["days_elapsed_mtd"])
        ws.cell(row=ri, column=M_DAYS).value    = _n(d["total_days_in_month"])

    last_data = DATA_START_ROW - 1 + len(rows)
    _write_totals(ws, last_data + 1, last_data,
        sum_blue   = {M_MTD, M_PRIOR, M_L30, M_FCST, M_G_MTD, M_G_P, M_G_L, M_G_F, M_B_MTD, M_B_P, M_B_L, M_B_F},
        sum_purple = {M_V1_MTD, M_V1_P, M_V1_L, M_V2_MTD, M_V2_P, M_V2_L},
        sum_green  = {M_O_MTD, M_O_P, M_O_L, M_O_F})
    ws.column_dimensions[get_column_letter(M_ELAPSED)].hidden = True
    ws.column_dimensions[get_column_letter(M_DAYS)].hidden    = True

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 – FW
# ─────────────────────────────────────────────────────────────────────────────
def build_fw_sheet(wb, rows):
    ws = wb.create_sheet("FW Partners")
    ws.freeze_panes = "C6"
    today = datetime.today()

    _write_title(ws,
        f"FW Partner Leads Report  –  {today.strftime('%B %d, %Y')}",
        "Forecast: MTD ÷ Days elapsed × Days in month  |  Source identified via LEAD_DATA.PSRC = 'FW'",
        F_VIS_COLS, title_fill=DARK_ORANGE)

    # Row 3: single section banner
    ws.merge_cells(f"A3:{get_column_letter(F_VIS_COLS)}3")
    ws["A3"] = "◀  PARTNER PERFORMANCE"
    ws["A3"].font = _font(bold=True, size=10); ws["A3"].fill = _fill(DARK_ORANGE)
    ws["A3"].alignment = _center(); ws.row_dimensions[3].height = 20

    # Row 4: blank filler row
    for ci in range(1, F_VIS_COLS + 1):
        ws.cell(row=4, column=ci).fill = _fill(MID_ORANGE)
        ws.cell(row=4, column=ci).border = _border()
    ws.row_dimensions[4].height = 8

    # Row 5: column headers
    col_specs = [
        (F_PID,   "Partner\nID",             12, MID_ORANGE, WHITE, None),
        (F_PNAME, "Partner Name",            28, MID_ORANGE, WHITE, None),
        (F_MTD,   "Leads\nMTD",              12, MID_ORANGE, WHITE, None),
        (F_PRIOR, "Leads\nPrior Month",      14, MID_ORANGE, WHITE, None),
        (F_L30,   "Leads\nLast 30 Days",     14, MID_ORANGE, WHITE, None),
        (F_FCST,  "Forecast\n(Full Month)",  14, MID_ORANGE, WHITE, None),
        (F_CPL,   "Cost Per\nLead",          12, MID_ORANGE, WHITE, None),
        (F_SCRUB, "Permissible\nScrub Rate", 13, MID_ORANGE, WHITE, None),
    ]
    for ci, lbl, w, fill_c, font_c, thick_c in col_specs:
        cell = ws.cell(row=5, column=ci, value=lbl)
        cell.font = _font(bold=True, color=font_c, size=9); cell.fill = _fill(fill_c)
        cell.alignment = _center(); cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.cell(row=5, column=F_ELAPSED, value="Days Elapsed")
    ws.cell(row=5, column=F_DAYS,    value="Days in Month")
    ws.row_dimensions[5].height = 36

    # Data rows
    for ri, row in enumerate(rows, start=DATA_START_ROW):
        d  = {col: row[i] for i, col in enumerate(FW_COLUMNS)}
        bg = LIGHT_ORANGE if ri % 2 == 0 else WHITE

        def put(ci, value, fmt=None, fill_c=None, font_c="000000", bold=False):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.fill = _fill(fill_c if fill_c is not None else bg)
            cell.font = _font(color=font_c, size=10, bold=bold)
            cell.alignment = _center(); cell.border = _border()
            if fmt: cell.number_format = fmt

        put(F_PID,   d["partner_id"])
        put(F_PNAME, d["partner_name"]); ws.cell(row=ri, column=F_PNAME).alignment = _left()
        put(F_MTD,   _n(d["leads_mtd"]),         "#,##0")
        put(F_PRIOR, _n(d["leads_prior_month"]),  "#,##0")
        put(F_L30,   _n(d["leads_last_30"]),      "#,##0")
        put(F_FCST,  _fcst(F_MTD, F_ELAPSED, F_DAYS, ri), "#,##0", font_c=FCST_BLUE, bold=True)
        put(F_CPL,   _n(d["cost_per_lead"]),       '"$"#,##0.00')
        put(F_SCRUB, _n(d["scrub_rate"]),          "0.0%")
        ws.cell(row=ri, column=F_ELAPSED).value = _n(d["days_elapsed_mtd"])
        ws.cell(row=ri, column=F_DAYS).value    = _n(d["total_days_in_month"])

    last_data = DATA_START_ROW - 1 + len(rows)
    _write_totals(ws, last_data + 1, last_data,
        sum_blue={F_MTD, F_PRIOR, F_L30, F_FCST}, sum_purple=set(), sum_green=set())
    ws.column_dimensions[get_column_letter(F_ELAPSED)].hidden = True
    ws.column_dimensions[get_column_letter(F_DAYS)].hidden    = True

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 – Rollup
# ─────────────────────────────────────────────────────────────────────────────
def build_rollup_sheet(wb, mil_rows, fw_rows):
    ws = wb.create_sheet("Rollup")
    ws.freeze_panes = "A5"
    today = datetime.today()

    # Column layout: Partner Name | BU | MTD | Prior Month | Last 30 | Forecast | CPL | Scrub Rate
    R_NAME=1; R_BU=2; R_MTD=3; R_PRIOR=4; R_L30=5; R_FCST=6; R_CPL=7; R_SCRUB=8
    R_ELAPSED=9; R_DAYS=10; R_VIS=8

    _write_title(ws,
        f"Partner Leads Rollup  –  {today.strftime('%B %d, %Y')}",
        "Combined view of MIL and FW partners. For source and version breakdown see individual tabs.",
        R_VIS, title_fill=DARK_BLUE)

    ws.merge_cells(f"A3:{get_column_letter(R_VIS)}3")
    ws["A3"] = "◀  ALL PARTNERS — COMBINED ROLLUP"
    ws["A3"].font = _font(bold=True, size=10); ws["A3"].fill = _fill(DARK_BLUE)
    ws["A3"].alignment = _center(); ws.row_dimensions[3].height = 20

    col_specs = [
        (R_NAME,  "Partner Name",            28, MID_BLUE, WHITE),
        (R_BU,    "BU",                       8, MID_BLUE, WHITE),
        (R_MTD,   "Leads\nMTD",              12, MID_BLUE, WHITE),
        (R_PRIOR, "Leads\nPrior Month",      14, MID_BLUE, WHITE),
        (R_L30,   "Leads\nLast 30 Days",     14, MID_BLUE, WHITE),
        (R_FCST,  "Forecast\n(Full Month)",  14, MID_BLUE, WHITE),
        (R_CPL,   "Cost Per\nLead",          12, MID_BLUE, WHITE),
        (R_SCRUB, "Permissible\nScrub Rate", 13, MID_BLUE, WHITE),
    ]
    for ci, lbl, w, fill_c, font_c in col_specs:
        cell = ws.cell(row=4, column=ci, value=lbl)
        cell.font = _font(bold=True, color=font_c, size=9); cell.fill = _fill(fill_c)
        cell.alignment = _center(); cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.cell(row=4, column=R_ELAPSED, value="Days Elapsed")
    ws.cell(row=4, column=R_DAYS,    value="Days in Month")
    ws.row_dimensions[4].height = 36

    # Combine MIL and FW rows, tagging each with BU
    all_rows = []
    for row in mil_rows:
        d = {col: row[i] for i, col in enumerate(MIL_COLUMNS)}
        all_rows.append(("MIL", d["partner_name"], _n(d["leads_mtd"]),
                         _n(d["leads_prior_month"]), _n(d["leads_last_30"]),
                         _n(d["cost_per_lead"]), _n(d["scrub_rate"]),
                         _n(d["days_elapsed_mtd"]), _n(d["total_days_in_month"])))
    for row in fw_rows:
        d = {col: row[i] for i, col in enumerate(FW_COLUMNS)}
        all_rows.append(("FW", d["partner_name"], _n(d["leads_mtd"]),
                         _n(d["leads_prior_month"]), _n(d["leads_last_30"]),
                         _n(d["cost_per_lead"]), _n(d["scrub_rate"]),
                         _n(d["days_elapsed_mtd"]), _n(d["total_days_in_month"])))
    all_rows.sort(key=lambda x: x[1])  # sort by partner name

    for ri, (bu, name, mtd, prior, l30, cpl, scrub, elapsed, days_tot) in enumerate(all_rows, start=5):
        bg     = LIGHT_BLUE if bu == "MIL" and ri % 2 == 0 else \
                 "EBF3FB"   if bu == "MIL" else \
                 LIGHT_ORANGE if ri % 2 == 0 else WHITE
        bu_color = DARK_BLUE if bu == "MIL" else DARK_ORANGE

        def put(ci, value, fmt=None, font_c="000000", bold=False):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.fill = _fill(bg); cell.font = _font(color=font_c, size=10, bold=bold)
            cell.alignment = _center(); cell.border = _border()
            if fmt: cell.number_format = fmt

        put(R_NAME,  name);  ws.cell(row=ri, column=R_NAME).alignment = _left()
        put(R_BU,    bu,    font_c=bu_color, bold=True)
        put(R_MTD,   mtd,   "#,##0")
        put(R_PRIOR, prior, "#,##0")
        put(R_L30,   l30,   "#,##0")
        el = get_column_letter(R_ELAPSED); tot = get_column_letter(R_DAYS)
        mtd_l = get_column_letter(R_MTD)
        ws.cell(row=ri, column=R_FCST).value = f'=IFERROR(ROUND({mtd_l}{ri}/MAX({el}{ri},1)*{tot}{ri},0),"")'
        ws.cell(row=ri, column=R_FCST).number_format = "#,##0"
        ws.cell(row=ri, column=R_FCST).font = _font(color=FCST_BLUE, size=10, bold=True)
        ws.cell(row=ri, column=R_FCST).fill = _fill(bg)
        ws.cell(row=ri, column=R_FCST).alignment = _center()
        ws.cell(row=ri, column=R_FCST).border = _border()
        put(R_CPL,   cpl,   '"$"#,##0.00')
        put(R_SCRUB, scrub, "0.0%")
        ws.cell(row=ri, column=R_ELAPSED).value = elapsed
        ws.cell(row=ri, column=R_DAYS).value    = days_tot

    last_data = 4 + len(all_rows)
    tr = last_data + 1
    ws.row_dimensions[tr].height = 20
    for ci in range(1, R_VIS + 1):
        cell = ws.cell(row=tr, column=ci)
        cell.fill = _fill(DARK_BLUE); cell.border = _border()
        if ci == R_NAME:
            cell.value = "TOTALS"; cell.font = _font(bold=True, size=10); cell.alignment = _center()
        elif ci in {R_MTD, R_PRIOR, R_L30, R_FCST}:
            col_ltr = get_column_letter(ci)
            cell.value = f"=SUM({col_ltr}5:{col_ltr}{last_data})"
            cell.number_format = "#,##0"
            cell.font = _font(bold=True, color=ACCENT_BLUE, size=10); cell.alignment = _center()

    ws.column_dimensions[get_column_letter(R_ELAPSED)].hidden = True
    ws.column_dimensions[get_column_letter(R_DAYS)].hidden    = True

# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def build_workbook(mil_rows, fw_rows):
    wb = openpyxl.Workbook()
    build_mil_sheet(wb, mil_rows)
    build_fw_sheet(wb, fw_rows)
    build_rollup_sheet(wb, mil_rows, fw_rows)
    wb.save(OUTPUT_FILE)
    print(f"✅  Report saved:  {OUTPUT_FILE}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Partner Leads Report Generator")
    parser.add_argument("--csv", action="store_true",
                        help="Load from CSV exports instead of Oracle")
    args = parser.parse_args()

    if args.csv:
        print("📄 Loading from CSV files...")
        mil_data = fetch_csv("mil_results.csv", MIL_COLUMNS)
        fw_data  = fetch_csv("fw_results.csv",  FW_COLUMNS)
    else:
        print("🔌 Connecting to Oracle...")
        mil_data = fetch_oracle(MIL_QUERY_FILE)
        fw_data  = fetch_oracle(FW_QUERY_FILE)

    build_workbook(mil_data, fw_data)
