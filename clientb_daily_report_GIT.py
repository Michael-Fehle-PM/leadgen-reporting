"""
fw_daily_report.py
==================
Produces the FW Daily Leads workbook — one tab per date in the range,
each tab showing:

  LEFT SUMMARY (cols A-D):
      Date | Leads Collected (UTC→ET) | Leads Sent | Delta (Sent - Collected)

  RIGHT DETAIL (cols F-G):
      Partner ID | Partner Name | Leads Sent on that date

DATE RANGE
----------
Defaults to May 1 of the current year through yesterday.
Override via command line:
    python fw_daily_report.py --start 2026-05-01 --end 2026-05-10

UTC OFFSET
----------
Automatically determined: EDT (UTC-4) Mar–Nov, EST (UTC-5) Nov–Mar.
Override via command line:
    python fw_daily_report.py --utc-offset -5

HOW TO RUN
----------
    python fw_daily_report.py
    python fw_daily_report.py --start 2026-05-01 --end 2026-05-10
"""

import argparse
import os
import sys
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Oracle connection ─────────────────────────────────────────────────────────
DB_USER     = "withheld"
DB_PASSWORD = "withheld"
DB_DSN      = "withheld"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
QUERY_FILE      = os.path.join(SCRIPT_DIR, "query_fw_daily.sql")

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_ORANGE  = "7B3F00"
MID_ORANGE   = "C55A11"
LIGHT_ORANGE = "FCE4D6"
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "D6E4F0"
RED          = "C00000"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F2F2F2"

# ── Style helpers ─────────────────────────────────────────────────────────────
def _font(bold=False, color=WHITE, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_left(left_color):
    return Border(left=Side(style="medium", color=left_color),
                  right=Side(style="thin",   color="BFBFBF"),
                  top=Side(style="thin",     color="BFBFBF"),
                  bottom=Side(style="thin",  color="BFBFBF"))

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _n(val):
    try:    return int(float(val))
    except: return 0

# ── UTC offset: auto-detect EDT/EST based on US DST rules ────────────────────
def get_utc_offset(for_date=None):
    """
    Returns the ET UTC offset for a given date.
    US DST: starts 2nd Sunday in March, ends 1st Sunday in November.
    EDT = UTC-4, EST = UTC-5.
    """
    if for_date is None:
        for_date = date.today()

    year = for_date.year

    # 2nd Sunday in March
    march1   = date(year, 3, 1)
    dst_start = march1 + timedelta(days=(6 - march1.weekday()) % 7 + 7)

    # 1st Sunday in November
    nov1     = date(year, 11, 1)
    dst_end  = nov1 + timedelta(days=(6 - nov1.weekday()) % 7)

    if dst_start <= for_date < dst_end:
        return -4   # EDT
    return -5       # EST

# ── Data fetching ─────────────────────────────────────────────────────────────
def fetch_data(start_dt, end_dt, utc_offset):
    try:
        import oracledb
    except ImportError:
        sys.exit("oracledb not installed.  Run:  pip install oracledb")

    with open(QUERY_FILE) as fh:
        sql = fh.read()

    conn = oracledb.connect(user=DB_USER, password=DB_PASSWORD, dsn=DB_DSN)
    cur  = conn.cursor()
    cur.execute(sql, {
        "start_date":  start_dt,
        "end_date":    end_dt,
        "utc_offset":  utc_offset,
    })
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows

# ── Group rows by date ────────────────────────────────────────────────────────
def group_by_date(rows):
    """
    Returns dict: { date_obj: { 'collected': n, 'sent_total': n,
                                 'partners': [(pid, name, sent, cpl, scrub), ...] } }
    """
    grouped = {}
    for row in rows:
        report_date, partner_id, partner_name, leads_sent, \
            cost_per_lead, scrub_rate, leads_collected, leads_sent_total = row

        if hasattr(report_date, 'date'):
            report_date = report_date.date()

        if report_date not in grouped:
            grouped[report_date] = {
                "collected":   _n(leads_collected),
                "sent_total":  _n(leads_sent_total),
                "partners":    [],
            }
        grouped[report_date]["partners"].append((
            _n(partner_id), str(partner_name), _n(leads_sent),
            float(cost_per_lead or 0), float(scrub_rate or 0),
        ))

    return dict(sorted(grouped.items()))

# ── Build one worksheet per date ──────────────────────────────────────────────
def build_sheet(wb, report_date, day_data, first_sheet=False):
    tab_label = report_date.strftime("%b %d").replace(" 0", " ")  # "May 1", "May 14"

    if first_sheet:
        ws = wb.active
        ws.title = tab_label
    else:
        ws = wb.create_sheet(title=tab_label)

    collected  = day_data["collected"]
    sent_total = day_data["sent_total"]
    delta      = sent_total - collected
    partners   = day_data["partners"]

    date_label = report_date.strftime("%B %-d %Y") if sys.platform != "win32" \
                 else report_date.strftime("%B %#d %Y")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14  # Date
    ws.column_dimensions["B"].width = 13  # Collected
    ws.column_dimensions["C"].width = 13  # Sent
    ws.column_dimensions["D"].width = 10  # Delta
    ws.column_dimensions["E"].width = 3   # Spacer
    ws.column_dimensions["F"].width = 12  # Partner ID
    ws.column_dimensions["G"].width = 30  # Partner Name
    ws.column_dimensions["H"].width = 13  # Leads Sent
    ws.column_dimensions["I"].width = 12  # CPL
    ws.column_dimensions["J"].width = 12  # Scrub Rate
    ws.column_dimensions["K"].width = 15  # Max Revenue
    ws.column_dimensions["L"].width = 15  # Min Revenue

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"] = f"FW Daily Leads  –  {date_label}"
    ws["A1"].font      = _font(bold=True, size=13)
    ws["A1"].fill      = _fill(DARK_ORANGE)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 24

    # ── Row 2: Section headers ────────────────────────────────────────────────
    ws.merge_cells("B2:D2")
    ws["B2"] = "LEADS"
    ws["B2"].font      = _font(bold=True, size=10)
    ws["B2"].fill      = _fill(MID_ORANGE)
    ws["B2"].alignment = _center()
    ws["B2"].border    = _border()
    for ci in [1, 5]:
        ws.cell(row=2, column=ci).fill = _fill(MID_ORANGE)

    ws.merge_cells("F2:L2")
    ws["F2"] = "PARTNER BREAKDOWN"
    ws["F2"].font      = _font(bold=True, size=10)
    ws["F2"].fill      = _fill(MID_BLUE)
    ws["F2"].alignment = _center()
    ws["F2"].border    = _thick_left(DARK_BLUE)
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    summary_headers = [
        (1,  "",              DARK_ORANGE, WHITE, None),
        (2,  "Collected",     MID_ORANGE,  WHITE, None),
        (3,  "Sent",          MID_ORANGE,  WHITE, None),
        (4,  "Δ",             MID_ORANGE,  WHITE, None),
        (5,  "",              MID_ORANGE,  WHITE, None),
        (6,  "Partner ID",    MID_BLUE,    WHITE, DARK_BLUE),
        (7,  "Partner Name",  MID_BLUE,    WHITE, None),
        (8,  "Leads Sent",    MID_BLUE,    WHITE, None),
        (9,  "Cost Per Lead", MID_BLUE,    WHITE, None),
        (10, "Scrub Rate",    MID_BLUE,    WHITE, None),
        (11, "Max Revenue",   MID_BLUE,    WHITE, None),
        (12, "Min Revenue",   MID_BLUE,    WHITE, None),
    ]
    for ci, lbl, fill_c, font_c, thick_c in summary_headers:
        cell = ws.cell(row=3, column=ci, value=lbl)
        cell.font      = _font(bold=True, color=font_c, size=10)
        cell.fill      = _fill(fill_c)
        cell.alignment = _center()
        cell.border    = _thick_left(thick_c) if thick_c else _border()
    ws.row_dimensions[3].height = 20

    # ── Row 4: Summary data row ───────────────────────────────────────────────
    delta_color = RED if delta < 0 else "375623"

    def put_sum(ci, value, fmt=None, font_c="000000", bold=False, fill_c=LIGHT_ORANGE, bdr=None):
        cell = ws.cell(row=4, column=ci, value=value)
        cell.fill      = _fill(fill_c)
        cell.font      = _font(color=font_c, size=10, bold=bold)
        cell.alignment = _center()
        cell.border    = bdr if bdr else _border()
        if fmt: cell.number_format = fmt

    put_sum(1, date_label, font_c=DARK_ORANGE, bold=True)
    ws.cell(row=4, column=1).alignment = _left()
    put_sum(2, collected,  "#,##0")
    put_sum(3, sent_total, "#,##0")
    put_sum(4, delta,      "+#,##0;-#,##0;0", font_c=delta_color, bold=True)
    put_sum(5, "",         fill_c=WHITE)  # spacer
    # cols 6-12 are partner detail — leave row 4 blank for non-first partners
    for ci in range(6, 13):
        ws.cell(row=4, column=ci).fill   = _fill(WHITE)
        ws.cell(row=4, column=ci).border = _border()

    # ── Rows 4+: Partner detail (right side, starting row 4) ─────────────────
    for pi, (pid, pname, psent, cpl, scrub) in enumerate(partners, start=4):
        bg = LIGHT_BLUE if pi % 2 == 0 else WHITE

        def put_p(ci, value, fmt=None, font_c="000000", bold=False, bdr=None):
            cell = ws.cell(row=pi, column=ci, value=value)
            cell.fill      = _fill(bg)
            cell.font      = _font(color=font_c, size=10, bold=bold)
            cell.alignment = _center()
            cell.border    = bdr if bdr else _border()
            if fmt: cell.number_format = fmt

        # Max Revenue = Leads Sent × CPL
        # Min Revenue = Leads Sent × (1 - Scrub Rate) × CPL
        sent_col  = get_column_letter(8)
        cpl_col   = get_column_letter(9)
        scrub_col = get_column_letter(10)

        put_p(6,  pid,   bdr=_thick_left(DARK_BLUE))
        put_p(7,  pname, font_c=DARK_BLUE)
        ws.cell(row=pi, column=7).alignment = _left()
        put_p(8,  psent, "#,##0",      bold=(psent > 0))
        put_p(9,  cpl,   '"$"#,##0.00')
        put_p(10, scrub, "0.0%")
        put_p(11, f"={sent_col}{pi}*{cpl_col}{pi}",
              '"$"#,##0.00', font_c=DARK_BLUE)
        put_p(12, f"={sent_col}{pi}*(1-{scrub_col}{pi})*{cpl_col}{pi}",
              '"$"#,##0.00', font_c="375623")

    # ── Partner totals row ────────────────────────────────────────────────────
    last_partner_row = 3 + len(partners)
    tr = last_partner_row + 1
    for ci in [6, 7, 8, 11, 12]:
        cell = ws.cell(row=tr, column=ci)
        cell.fill      = _fill(DARK_BLUE)
        cell.border    = _border()
        if ci == 7:
            cell.value     = "TOTAL"
            cell.font      = _font(bold=True, size=10)
            cell.alignment = _center()
        else:
            col_ltr            = get_column_letter(ci)
            cell.value         = f"=SUM({col_ltr}4:{col_ltr}{last_partner_row})"
            cell.number_format = "#,##0.00" if ci in [11, 12] else "#,##0"
            cell.font          = _font(bold=True, color="00B0F0", size=10)
            cell.alignment     = _center()
    # Fill remaining total cells
    for ci in [6, 9, 10]:
        cell           = ws.cell(row=tr, column=ci)
        cell.fill      = _fill(DARK_BLUE)
        cell.border    = _border()
    ws.row_dimensions[tr].height = 18

# ── Output path ───────────────────────────────────────────────────────────────
def _output_path(start_dt, end_dt):
    base_dir  = SCRIPT_DIR
    stamp     = f"{start_dt.strftime('%Y%m%d')}_to_{end_dt.strftime('%Y%m%d')}"
    path = os.path.join(base_dir, f"FW_Daily_Leads_{stamp}.xlsx")
    if not os.path.exists(path):
        return path
    counter = 1
    while True:
        path = os.path.join(base_dir, f"FW_Daily_Leads_{stamp}_{counter}.xlsx")
        if not os.path.exists(path):
            return path
        counter += 1

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="FW Daily Leads Report")
    parser.add_argument("--start",      default=None,
                        help="Start date YYYY-MM-DD (default: May 1 of current year)")
    parser.add_argument("--end",        default=None,
                        help="End date YYYY-MM-DD (default: yesterday)")
    parser.add_argument("--utc-offset", type=int, default=None,
                        help="UTC offset for ET conversion e.g. -4 or -5 (default: auto)")
    args = parser.parse_args()

    today     = date.today()
    start_dt  = date(today.year, 5, 1) if not args.start \
                else datetime.strptime(args.start, "%Y-%m-%d").date()
    end_dt    = today - timedelta(days=1) if not args.end \
                else datetime.strptime(args.end, "%Y-%m-%d").date()
    utc_offset = args.utc_offset if args.utc_offset is not None \
                 else get_utc_offset(start_dt)

    print(f"📅 Date range : {start_dt} → {end_dt}")
    print(f"🕐 UTC offset : {utc_offset} (ET)")
    print("🔌 Connecting to Oracle...")

    rows    = fetch_data(start_dt, end_dt, utc_offset)
    grouped = group_by_date(rows)

    if not grouped:
        print("⚠️  No data returned for the requested date range.")
        sys.exit(0)

    print(f"📊 Building workbook ({len(grouped)} tabs)...")
    wb = openpyxl.Workbook()

    for i, (report_date, day_data) in enumerate(grouped.items()):
        build_sheet(wb, report_date, day_data, first_sheet=(i == 0))
        print(f"   ✓ {report_date.strftime('%B %d, %Y')}  "
              f"({len(day_data['partners'])} partners, "
              f"{day_data['sent_total']} sent, "
              f"{day_data['collected']} collected)")

    output_file = _output_path(start_dt, end_dt)
    wb.save(output_file)
    print(f"\n✅  Report saved:  {output_file}")

if __name__ == "__main__":
    main()
